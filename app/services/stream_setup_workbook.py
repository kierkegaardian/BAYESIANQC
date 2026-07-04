from __future__ import annotations

from io import BytesIO
from typing import Any

from pydantic import ValidationError

from app.stream_setup_models import StreamSetupBatchIn, StreamSetupIn, StreamSetupPreviewRow

STREAM_HEADERS = [
    "stream_id",
    "site",
    "lab_bench",
    "instrument_name",
    "instrument_manufacturer",
    "instrument_model",
    "method_name",
    "method_technique",
    "parameter_name",
    "units",
    "material_name",
    "material_manufacturer",
    "matrix",
    "qc_level",
    "control_material_lot",
    "target_value",
    "sigma",
    "warning_limit_sd",
    "action_limit_sd",
    "min_value",
    "max_value",
    "effective_from",
    "config_reason",
]
PRIOR_HEADERS = ["stream_id", "prior_mu0", "prior_kappa0", "prior_alpha0", "prior_beta0", "prior_effective_from"]
KIOSK_HEADERS = [
    "stream_id",
    "kiosk_slug",
    "kiosk_label",
    "panel_title",
    "panel_start",
    "panel_end",
    "panel_window_label",
    "mode",
]
TEMPLATE_SHEETS = {
    "Locations": ["site", "lab_bench"],
    "Instruments": ["instrument_name", "instrument_manufacturer", "instrument_model", "site", "lab_bench"],
    "Methods": ["instrument_name", "method_name", "method_technique"],
    "Parameters": ["instrument_name", "method_name", "parameter_name", "units"],
    "Materials": ["material_name", "material_manufacturer", "matrix", "qc_level", "control_material_lot"],
    "Streams": STREAM_HEADERS,
    "Priors": PRIOR_HEADERS,
    "KioskAssignments": KIOSK_HEADERS,
}


def workbook_template_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    first = True
    for title, headers in TEMPLATE_SHEETS.items():
        sheet = workbook.active if first else workbook.create_sheet(title)
        if sheet is None:
            raise RuntimeError("Workbook did not create an active worksheet")
        first = False
        sheet.title = title
        sheet.append(headers)
        sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _rows_by_stream(sheet: Any) -> dict[str, dict[str, object]]:
    rows: dict[str, dict[str, object]] = {}
    if sheet is None:
        return rows
    headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        data = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and value not in ("", None)}
        stream_id = str(data.get("stream_id") or "").strip()
        if stream_id:
            rows[stream_id] = data
    return rows


def _stream_rows(sheet: Any) -> list[tuple[int, dict[str, object]]]:
    if sheet is None:
        raise ValueError("Workbook must include a Streams sheet")
    headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[tuple[int, dict[str, object]]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and value not in ("", None)}
        if data:
            rows.append((row_number, data))
    return rows


def parse_workbook(data: bytes) -> tuple[StreamSetupBatchIn, list[StreamSetupPreviewRow]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(data), data_only=True)
    prior_rows = _rows_by_stream(workbook["Priors"] if "Priors" in workbook.sheetnames else None)
    kiosk_rows = _rows_by_stream(workbook["KioskAssignments"] if "KioskAssignments" in workbook.sheetnames else None)
    parsed: list[StreamSetupIn] = []
    invalid: list[StreamSetupPreviewRow] = []
    for row_number, row in _stream_rows(workbook["Streams"] if "Streams" in workbook.sheetnames else None):
        stream_id = str(row.get("stream_id") or "").strip()
        merged = {**row, **prior_rows.get(stream_id, {})}
        kiosk = kiosk_rows.get(stream_id)
        if kiosk:
            merged["kiosk"] = {key: value for key, value in kiosk.items() if key != "stream_id"}
        try:
            parsed.append(StreamSetupIn.model_validate(merged))
        except ValidationError as exc:
            invalid.append(
                StreamSetupPreviewRow(
                    row=row_number,
                    stream_id=stream_id or f"row-{row_number}",
                    valid=False,
                    errors=[f"{'.'.join(str(part) for part in error['loc'])}: {error['msg']}" for error in exc.errors()],
                    actions=[],
                    canonical=None,
                )
            )
    return StreamSetupBatchIn(rows=parsed), invalid
