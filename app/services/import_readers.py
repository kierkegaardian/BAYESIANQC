from __future__ import annotations

import csv
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

from app.import_db_models import ParserProfile
from app.import_models import ImportRowType, ParserProfileType


@dataclass(frozen=True)
class SourceRow:
    row_number: int
    row_type: ImportRowType
    raw: dict[str, Any]
    errors: list[str]


def _text(data: bytes, encoding: str) -> str:
    try:
        return data.decode(encoding)
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=422, detail=f"File is not valid {encoding} text") from exc


def _delimiter(profile: ParserProfile, filename: str) -> str:
    configured = profile.config.get("delimiter")
    if isinstance(configured, str) and configured:
        return "\t" if configured == "\\t" else configured
    return "\t" if Path(filename).suffix.lower() in {".txt", ".dat"} else ","


def _normalize_raw(raw: dict[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for key, value in raw.items():
        if key is None:
            continue
        label = str(key).strip()
        if not label:
            continue
        cleaned[label] = value.strip() if isinstance(value, str) else value
    return cleaned


def _row_is_blank(raw: dict[str, Any]) -> bool:
    return all(value in (None, "") for value in raw.values())


def _matches_expected(raw: dict[str, Any], profile: ParserProfile) -> bool:
    expected = profile.config.get("expected_tests")
    if not isinstance(expected, list) or not expected:
        return True
    columns = profile.config.get("columns")
    mapped_analyte = columns.get("analyte") if isinstance(columns, dict) else None
    analyte_column = str(profile.config.get("analyte_column") or mapped_analyte or "")
    token = str(raw.get(analyte_column, "")).strip().lower() if analyte_column else " ".join(map(str, raw.values())).lower()
    aliases: set[str] = set()
    for item in expected:
        if isinstance(item, str):
            aliases.add(item.lower())
        elif isinstance(item, dict):
            name = item.get("name")
            if isinstance(name, str):
                aliases.add(name.lower())
            values = item.get("aliases")
            if isinstance(values, list):
                aliases.update(str(value).lower() for value in values)
    return any(alias and alias in token for alias in aliases)


def _classify(raw: dict[str, Any], profile: ParserProfile) -> ImportRowType:
    if _row_is_blank(raw):
        return ImportRowType.IGNORED
    row_type_column = profile.config.get("row_type_column")
    if isinstance(row_type_column, str):
        value = str(raw.get(row_type_column, "")).strip().lower()
        configured = profile.config.get("row_type_values")
        if isinstance(configured, dict):
            for row_type, tokens in configured.items():
                if isinstance(tokens, list) and value in {str(token).lower() for token in tokens}:
                    try:
                        return ImportRowType(row_type)
                    except ValueError:
                        return ImportRowType.PARSE_ERROR
    if profile.config.get("peak_table"):
        peak_columns = {"retention_time", "area", "height", "peak_name"}
        column_values = profile.config.get("columns", {})
        if isinstance(column_values, dict) and peak_columns.intersection(column_values):
            return ImportRowType.PEAK
    if profile.profile_type == ParserProfileType.INSTRUMENT_TABLE_DISCOVERY and not _matches_expected(raw, profile):
        return ImportRowType.IGNORED
    return ImportRowType.QC_RESULT


def _read_delimited(data: bytes, filename: str, profile: ParserProfile) -> list[SourceRow]:
    text = _text(data, str(profile.config.get("encoding") or "utf-8-sig"))
    reader = csv.DictReader(StringIO(text), delimiter=_delimiter(profile, filename))
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="Delimited file has no header row")
    rows: list[SourceRow] = []
    for row_number, raw in enumerate(reader, start=2):
        normalized = _normalize_raw(raw)
        rows.append(SourceRow(row_number, _classify(normalized, profile), normalized, []))
    return rows


def _read_table_discovery(data: bytes, filename: str, profile: ParserProfile) -> list[SourceRow]:
    text = _text(data, str(profile.config.get("encoding") or "utf-8-sig"))
    lines = text.splitlines()
    anchor = profile.config.get("table_start")
    start = 0
    if isinstance(anchor, str) and anchor:
        for idx, line in enumerate(lines):
            if anchor.lower() in line.lower():
                start = idx + 1
                break
    remaining = [line for line in lines[start:] if line.strip()]
    if not remaining:
        raise HTTPException(status_code=422, detail="No table found for instrument discovery profile")
    reader = csv.DictReader(StringIO("\n".join(remaining)), delimiter=_delimiter(profile, filename))
    if not reader.fieldnames:
        raise HTTPException(status_code=422, detail="Discovered table has no header row")
    rows: list[SourceRow] = []
    for offset, raw in enumerate(reader, start=start + 2):
        normalized = _normalize_raw(raw)
        rows.append(SourceRow(offset, _classify(normalized, profile), normalized, []))
    return rows


def _read_xlsx(data: bytes, profile: ParserProfile) -> list[SourceRow]:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet_name = profile.config.get("sheet")
    sheet = workbook[str(sheet_name)] if isinstance(sheet_name, str) and sheet_name in workbook.sheetnames else workbook.active
    if sheet is None:
        raise HTTPException(status_code=422, detail="XLSX file has no active sheet")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="XLSX file has no rows")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed: list[SourceRow] = []
    for index, values in enumerate(rows[1:], start=2):
        raw = _normalize_raw(dict(zip(headers, values)))
        parsed.append(SourceRow(index, _classify(raw, profile), raw, []))
    return parsed


def _xml_value(element: ET.Element, path: str) -> str | None:
    if path.startswith("@"):
        return element.attrib.get(path[1:])
    child = element.find(path)
    return child.text.strip() if child is not None and child.text is not None else None


def _read_xml(data: bytes, profile: ParserProfile) -> list[SourceRow]:
    try:
        root = ET.fromstring(data)
    except ET.ParseError as exc:
        raise HTTPException(status_code=422, detail="XML file is not well formed") from exc
    row_path = str(profile.config.get("rows_path") or profile.config.get("row_tag") or ".")
    elements = root.findall(row_path) if row_path != "." else [root]
    columns = profile.config.get("columns")
    if not isinstance(columns, dict) or not columns:
        raise HTTPException(status_code=422, detail="XML mapping profile requires columns")
    rows: list[SourceRow] = []
    for index, element in enumerate(elements, start=1):
        raw = {field: _xml_value(element, str(path)) for field, path in columns.items()}
        rows.append(SourceRow(index, _classify(raw, profile), raw, []))
    return rows


def read_source_rows(data: bytes, filename: str, profile: ParserProfile) -> list[SourceRow]:
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx(data, profile)
    if profile.profile_type == ParserProfileType.XML_MAPPING or ext == ".xml":
        return _read_xml(data, profile)
    if profile.profile_type == ParserProfileType.INSTRUMENT_TABLE_DISCOVERY:
        return _read_table_discovery(data, filename, profile)
    if ext in {".csv", ".txt", ".dat"}:
        return _read_delimited(data, filename, profile)
    raise HTTPException(status_code=422, detail="Unsupported import file type")
