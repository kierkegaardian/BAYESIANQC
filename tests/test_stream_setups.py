from __future__ import annotations

from collections.abc import AsyncIterator
from datetime import datetime, timedelta, timezone
from io import BytesIO

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, col, select

from app.db import get_engine
from app.db_models import AuditEntry, ControlMaterial, Instrument, KioskPanel, StreamConfig
from app.main import app

AUTH_HEADERS = {"X-API-Key": "local-dev-key"}


@pytest.fixture
async def client() -> AsyncIterator[httpx.AsyncClient]:
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        yield client


def _setup_payload(**overrides: object) -> dict[str, object]:
    payload: dict[str, object] = {
        "stream_id": "fuel-bench-a-sulfur-l1-lot-a",
        "site": "Refinery Site 1",
        "lab_bench": "Bench A",
        "instrument_name": "Sindie 7039",
        "instrument_manufacturer": "XOS",
        "instrument_model": "Sindie",
        "method_name": "ASTM D7039",
        "method_technique": "MWDXRF",
        "parameter_name": "Total Sulfur",
        "units": "ppm",
        "material_name": "Sulfur QC",
        "material_manufacturer": "VHG",
        "matrix": "Diesel",
        "qc_level": "Level 1",
        "control_material_lot": "SULF-LOT-A",
        "target_value": 12.0,
        "sigma": 0.5,
        "warning_limit_sd": 2.0,
        "action_limit_sd": 3.0,
        "min_value": 0.0,
        "max_value": 20.0,
        "prior_mu0": 12.0,
        "prior_kappa0": 1.0,
        "prior_alpha0": 2.0,
        "prior_beta0": 0.25,
        "kiosk": {
            "kiosk_slug": "fuel-bench-a",
            "kiosk_label": "Fuel Bench A",
            "panel_title": "Sulfur L1",
            "panel_start": "2026-03-01",
            "panel_end": "2026-03-03",
            "panel_window_label": "Mar 1-3, 2026",
        },
    }
    payload.update(overrides)
    return payload


@pytest.mark.anyio
async def test_stream_setup_preview_apply_and_backlog_defaults(client: httpx.AsyncClient) -> None:
    payload = {"rows": [_setup_payload()]}
    preview = await client.post("/stream-setups/preview", json=payload, headers=AUTH_HEADERS)
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["valid"] == 1
    assert {action["entity"]: action["action"] for action in body["rows"][0]["actions"]}["stream"] == "create"

    applied = await client.post("/stream-setups/apply", json=payload, headers=AUTH_HEADERS)
    assert applied.status_code == 200, applied.text
    row = applied.json()["rows"][0]
    assert row["stream"]["lab_bench"] == "Bench A"
    assert row["stream"]["control_material_id"] == row["control_material"]["id"]
    assert row["kiosk"]["slug"] == "fuel-bench-a"
    assert row["kiosk"]["panels"][0]["stream_id"] == "fuel-bench-a-sulfur-l1-lot-a"

    stream_filter = await client.get("/streams?site=Refinery%20Site%201&lab_bench=Bench%20A", headers=AUTH_HEADERS)
    assert stream_filter.status_code == 200
    assert [stream["stream_id"] for stream in stream_filter.json()] == ["fuel-bench-a-sulfur-l1-lot-a"]
    kiosk_filter = await client.get("/kiosks?site=Refinery%20Site%201&lab_bench=Bench%20A", headers=AUTH_HEADERS)
    assert kiosk_filter.status_code == 200
    assert [kiosk["slug"] for kiosk in kiosk_filter.json()] == ["fuel-bench-a"]
    empty_kiosks = await client.get("/kiosks?site=Other", headers=AUTH_HEADERS)
    assert empty_kiosks.status_code == 200
    assert empty_kiosks.json() == []

    due_at = (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
    backlog = await client.post(
        "/qc/backlog",
        json={"source": "requested", "stream_id": "fuel-bench-a-sulfur-l1-lot-a", "due_at": due_at},
        headers=AUTH_HEADERS,
    )
    assert backlog.status_code == 200, backlog.text
    assert backlog.json()["lab_bench"] == "Bench A"

    with Session(get_engine()) as session:
        instrument = session.exec(select(Instrument).where(Instrument.name == "Sindie 7039")).one()
        stream = session.exec(select(StreamConfig).where(StreamConfig.stream_id == "fuel-bench-a-sulfur-l1-lot-a")).one()
        material = session.exec(select(ControlMaterial).where(ControlMaterial.lot == "SULF-LOT-A")).one()
        panel = session.exec(select(KioskPanel).where(KioskPanel.stream_id == stream.stream_id)).one()
        audit = session.exec(select(AuditEntry).where(AuditEntry.action == "create_stream").order_by(col(AuditEntry.id).desc())).first()
        assert instrument.lab_bench == "Bench A"
        assert stream.control_material_id == material.id
        assert panel.title == "Sulfur L1"
        assert audit is not None


@pytest.mark.anyio
async def test_stream_setup_conflict_requires_config_reason(client: httpx.AsyncClient) -> None:
    payload = {"rows": [_setup_payload()]}
    assert (await client.post("/stream-setups/apply", json=payload, headers=AUTH_HEADERS)).status_code == 200

    changed = _setup_payload(target_value=13.0)
    preview = await client.post("/stream-setups/preview", json={"rows": [changed]}, headers=AUTH_HEADERS)
    assert preview.status_code == 200
    assert preview.json()["invalid"] == 1
    assert "config_reason" in preview.json()["rows"][0]["errors"][0]

    changed["config_reason"] = "new assigned value from material CoA"
    versioned = await client.post("/stream-setups/apply", json={"rows": [changed]}, headers=AUTH_HEADERS)
    assert versioned.status_code == 200, versioned.text
    assert versioned.json()["rows"][0]["stream"]["version"] == 2


@pytest.mark.anyio
async def test_kiosk_panel_endpoint_rejects_unknown_stream(client: httpx.AsyncClient) -> None:
    created = await client.post("/kiosks", json={"slug": "lab-wall", "label": "Lab Wall"}, headers=AUTH_HEADERS)
    assert created.status_code == 200

    rejected = await client.post(
        "/kiosks/lab-wall/panels",
        json={"stream_id": "missing-stream", "title": "Missing"},
        headers=AUTH_HEADERS,
    )
    assert rejected.status_code == 422


@pytest.mark.anyio
async def test_xlsx_template_and_import_preview(client: httpx.AsyncClient) -> None:
    template = await client.get("/stream-setups/template.xlsx", headers=AUTH_HEADERS)
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content), read_only=True)
    assert {"Locations", "Streams", "Priors", "KioskAssignments"} <= set(workbook.sheetnames)

    upload_book = Workbook()
    streams = upload_book.active
    assert streams is not None
    streams.title = "Streams"
    streams.append(
        [
            "stream_id",
            "instrument_name",
            "method_name",
            "parameter_name",
            "units",
            "material_name",
            "qc_level",
            "control_material_lot",
            "target_value",
            "sigma",
        ]
    )
    streams.append(["xlsx-stream", "Analyzer", "Method", "Parameter", "mg/L", "Control", "L1", "LOT-X", 1.0, 0.1])
    streams.append(["bad-row", "Analyzer", "Method", None, "mg/L", "Control", "L1", "LOT-X", 1.0, 0.1])
    priors = upload_book.create_sheet("Priors")
    priors.append(["stream_id", "prior_mu0", "prior_kappa0", "prior_alpha0", "prior_beta0"])
    priors.append(["xlsx-stream", 1.0, 1.0, 2.0, 0.01])
    data = BytesIO()
    upload_book.save(data)
    data.seek(0)

    preview = await client.post(
        "/stream-setups/import/preview",
        files={"file": ("setup.xlsx", data.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=AUTH_HEADERS,
    )
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["valid"] == 1
    assert body["invalid"] == 1
    assert {row["stream_id"]: row["valid"] for row in body["rows"]} == {"xlsx-stream": True, "bad-row": False}
